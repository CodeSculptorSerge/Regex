from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook


def dataframe_to_excel(df, result_path):
    wb = Workbook()  # Создаем новую книгу Excel
    ws = wb.active  # Активируем текущий лист
    ws.title = "Результаты сопоставления"  # Переименовываем лист
    
    # Добавляем данные DataFrame в лист Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)  # Заполняем ячейки данными

    # Автоматическая подгонка ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:  # Проверяем наличие значения в ячейке
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  # Задаем новую ширину столбца
        ws.column_dimensions[column].width = adjusted_width

    # Определение диапазона данных для таблицы
    data_range = f'A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}'

    # Создание объекта таблицы с таблицей Excel из всего диапазона данных
    table = Table(displayName="ResultsTable", ref=data_range)

    # Присваивание стиля таблице для лучшего отображения
    style = TableStyleInfo(name="TableStyleMedium22", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)  # Добавляем стилизованную таблицу в лист

    # Сохранение файла Excel
    wb.save(result_path)  # Сохраняем книгу по указанному пути
